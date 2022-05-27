# This is a sample Python script.


from openpyxl import load_workbook


def read_excel(exist_path):
    wb = load_workbook(exist_path)
    ws = wb["Sheet1"]

    # 找出第一个字符为“香”的单元格行号 作为区分每一次的节点
    row_note = []
    for i in range(1, ws.max_row):
        if ws.cell(i, 1).value is not None:
            if ws.cell(i, 1).value[0] == "香":
                row_note.append(i)
    row_note.append(ws.max_row + 1)  # 最后一次的右侧范围

    # 缺失点
    a_lose = [1, 29, 31]
    b_lose = [5, 6, 11]
    c_lose = [1, 2, 4, 5, 9, 10, 22, 29, 31, 32]

    abc_lose = [a_lose, b_lose, c_lose]

    for items in abc_lose:
        for j in range(len(items)):
            if items[j] < 10:
                items[j] = '0' + str(items[j])
            else:
                items[j] = str(items[j])

    # 去除缺失点后每一轮abc三类行号
    abc_list = []
    # 划分去除缺失点后的每一轮各自节点行号
    abc_note = []
    # 所使用控制点行号
    kps_row = []

    # 循环后列表变为[ [],[],[] ] 分别对应a、b、c
    for i in range(3):
        abc_list.append([])
        abc_note.append([1])
        kps_row.append([])

    # 循环每一次 分别找出15次每一次的A、B、C、K行号
    for i in range(len(row_note) - 1):
        for j in range(row_note[i], row_note[i + 1]):
            if ws.cell(j, 1).value is not None:
                if ws.cell(j, 1).value[0] == "A":
                    if ws.cell(j, 1).value[-2:] not in abc_lose[0]: # 判断是否缺失
                        abc_list[0].append(j)
                elif ws.cell(j, 1).value[0] == "B":
                    if ws.cell(j, 1).value[-2:] not in abc_lose[1]:
                        abc_list[1].append(j)
                elif ws.cell(j, 1).value[0] == "C":
                    if ws.cell(j, 1).value[-2:] not in abc_lose[2]:
                        abc_list[2].append(j)
                elif ws.cell(j, 1).value[0] == "K":
                    if ws.cell(j, 1).value[1] == "A":
                        abc_list[0].append(j)
                    elif ws.cell(j, 1).value[1] == "B":
                        abc_list[1].append(j)
                    elif ws.cell(j, 1).value[1] == "C":
                        abc_list[2].append(j)
        for k in range(3):
            abc_note[k].append(len(abc_list[k]) + 1)
            # 15次每完成一次将记录这一次中ABC分别有多少个 便于后续构建节点列表

    # 创建A、B、C三个表，并将abc_list里记录的按顺序排列到三个表中
    ws2 = wb.create_sheet("ASheet")
    ws3 = wb.create_sheet("BSheet")
    ws4 = wb.create_sheet("CSheet")
    ws_list = [ws2, ws3, ws4]

    # 将去除缺失点后的每一行复制赋值到ABC各表中
    for m in range(3):
        for i in range(len(abc_list[m])):
            # for j in range(1, ws.max_column):
            for j in range(1, 5):
                ws_list[m].cell(i + 1, j).value = ws.cell(abc_list[m][i], j).value

    x_col = "B"
    x_col_num = ord(x_col) - 64
    y_col = "C"
    y_col_num = ord(y_col) - 64
    z_col = "D"
    z_col_num = ord(z_col) - 64

    # 简化两点距离表示
    def distance_str(row1, row2):
        return "=SQRT((" + x_col + row1 + "-" + x_col + row2 + ")^2+(" + y_col + row1 + "-" + y_col + row2 + ")^2)"

    # 计算水平距离
    horizon_distance_col = "E"  # 插入到哪一列
    horizon_distance_col_num = ord(horizon_distance_col) - 64  # 列转数字，便于cell索引
    # 计算控制点距离
    kp_distance_col = "J"
    kp_distance_col_num = ord(kp_distance_col) - 64
    kp_list = ["KA01", "KB04", "KC03"]

    # 为控制点行号列表赋值 循环15次 每一次都记录三个控制点行号
    for m in range(3):
        for i in range(len(abc_note[m]) - 1):
            for j in range(abc_note[m][i], abc_note[m][i + 1]):
                if ws_list[m].cell(j, 1).value in kp_list:
                    kps_row[m].append(j)

    for m in range(3):
        for i in range(len(abc_note[m]) - 1):
            for j in range(abc_note[m][i], abc_note[m][i + 1] - 1):

                if ws_list[m].cell(j, 1).value[0] == "ABC"[m]:
                    # 与控制点距离
                    ws_list[m].cell(j, kp_distance_col_num).value = distance_str(str(j), str(kps_row[m][i]))

                if ws_list[m].cell(j, 1).value[0] == "ABC"[m] and ws_list[m].cell(j + 1, 1).value[0] == "ABC"[m]:
                    # 水平距离计算
                    ws_list[m].cell(j, horizon_distance_col_num).value = distance_str(str(j + 1), str(j))

    first_observations_note = []  # 第一次索引
    follow_observations_note = []  # 后续观测索引
    for m in range(3):
        first_observations_note.append(abc_note[m][:2])
        follow_observations_note.append(abc_note[m][1:])

    # 声明下沉列数
    sink_depth_col = "G"  # 插入到哪一列
    sink_depth_col_num = ord(sink_depth_col) - 64  # 列转数字，便于cell索引
    # 声明倾斜列数
    incline_col = "H"  # 插入到哪一列
    incline_col_num = ord(incline_col) - 64
    # 声明曲率列数
    curvature_col = "I"
    curvature_col_num = ord(curvature_col) - 64
    # 声明水平移动列数
    horizon_move_col = "K"
    horizon_move_col_num = ord(horizon_move_col) - 64
    # 声明水平变形列数
    horizon_shape_col = "F"
    horizon_shape_col_num = ord(horizon_shape_col) - 64

    # 从第二次开始循环，减去第一次
    for m in range(3):
        for i in range(len(follow_observations_note[m]) - 1):
            for j in range(follow_observations_note[m][i], follow_observations_note[m][i + 1]):
                first_j = j - follow_observations_note[m][i] + 1
                if ws_list[m].cell(j, 1).value[0] == "ABC"[m]:
                    # 计算下沉
                    ws_list[m].cell(j, sink_depth_col_num).value = \
                        "=" + z_col + str(first_j) + "-" + z_col + str(j)
                    # 水平移动
                    ws_list[m].cell(j, horizon_move_col_num).value = "=" + kp_distance_col + str(
                        j) + "-" + kp_distance_col + str(first_j)

                    if ws_list[m].cell(j + 1, 1).value[0] == "K":
                        break
                    else:
                        # 计算倾斜，倾斜=下沉差值/水平距离
                        ws_list[m].cell(j, incline_col_num).value = \
                            "=(" + sink_depth_col + str(j + 1) + "-" + sink_depth_col + str(j) + ")/" + \
                            horizon_distance_col + str(first_j)
                        # 水平变形
                        ws_list[m].cell(j, horizon_shape_col_num).value = "=(" + horizon_distance_col + str(
                            j) + "-" + horizon_distance_col + str(first_j) + ")/" + horizon_distance_col + str(first_j)
                        # 计算曲率
                        # 曲率= 倾斜差值*2 /距离(（n+1）-n)+距离(n-（n-1）)
                        if j > follow_observations_note[m][i]:
                            ws_list[m].cell(j, curvature_col_num).value = "=(" + incline_col + str(
                                j) + "-" + incline_col + str(j - 1) + ")*2/" + "(" + distance_str(str(j + 1), str(j))[1:] + \
                                                                          "+" + distance_str(str(j), str(j - 1))[1:] + ")"

    # 总结计算出的参数
    attributes = (horizon_distance_col_num, horizon_shape_col_num, sink_depth_col_num, incline_col_num,
                  curvature_col_num, kp_distance_col_num, horizon_move_col_num)
    att_str = ("horizon_distance", "horizon_shape", "sink_depth", "incline", "curvature", "kp_distance", "horizon_move")

    # 为ABC各声明7个表
    ws_list_att = []
    for m in range(3):
        ws_list_att.append([])
        for i in range(len(attributes)):
            ws_list_att[m].append(wb.create_sheet("ABC"[m] + "_" + att_str[i]))

    for m in range(3):
        # 此函数接收两行号作为参数，计算行号为row1和row2的两点距离
        def distance_value(row1, row2):
            diff_x = ws_list[m].cell(row1, x_col_num).value - ws_list[m].cell(row2, x_col_num).value
            diff_y = ws_list[m].cell(row1, y_col_num).value - ws_list[m].cell(row2, y_col_num).value
            distance = (diff_x ** 2 + diff_y ** 2) ** 0.5
            return distance

        for i in range(len(follow_observations_note[m]) - 1):
            incline_stay = 0
            for j in range(follow_observations_note[m][i], follow_observations_note[m][i + 1]):
                # 当行号为j时 对应的第一次的点的行号为first_j
                first_j = j - follow_observations_note[m][i] + 1
                if ws_list[m].cell(j, 1).value[0] == "ABC"[m]:
                    # 计算初始水平距离
                    horizon_distance_first = distance_value(first_j + 1, first_j)
                    # 计算初始与控制点距离
                    kp_distance_first = distance_value(first_j, kps_row[m][0])
                    # 计算下沉
                    sink_depth = ws_list[m].cell(first_j, z_col_num).value - ws_list[m].cell(j, z_col_num).value
                    # 添加序号
                    if i == 0:
                        ws_list_att[m][2].cell(first_j, i + 1).value = first_j
                        ws_list_att[m][5].cell(first_j, i + 1).value = first_j
                        ws_list_att[m][6].cell(first_j, i + 1).value = first_j

                    ws_list_att[m][2].cell(first_j, i + 2).value = sink_depth
                    # 计算与控制点距离
                    kp_distance = distance_value(j,kps_row[m][i])
                    ws_list_att[m][5].cell(first_j, i + 2).value = kp_distance
                    # 计算水平移动
                    horizon_move = kp_distance - kp_distance_first
                    ws_list_att[m][6].cell(first_j, i + 2).value = horizon_move
                    if ws_list[m].cell(j + 1, 1).value[0] == "K":
                        break
                    else:
                        # 计算水平距离
                        horizon_distance = distance_value(j+1,j)
                        ws_list_att[m][0].cell(first_j, i + 2).value = horizon_distance
                        # 增加序号
                        if i == 0:
                            ws_list_att[m][0].cell(first_j, i + 1).value = first_j
                            ws_list_att[m][1].cell(first_j, i + 1).value = first_j
                            ws_list_att[m][3].cell(first_j, i + 1).value = first_j
                        # 计算水平变形
                        horizon_shape = (horizon_distance - horizon_distance_first) / horizon_distance_first
                        ws_list_att[m][1].cell(first_j, i + 2).value = horizon_shape
                        # 计算倾斜
                        sink_depth_next = ws_list[m].cell(first_j + 1, 4).value - ws_list[m].cell(j + 1, 4).value
                        incline = (sink_depth_next - sink_depth) / horizon_distance_first
                        ws_list_att[m][3].cell(first_j, i + 2).value = incline

                        if j > follow_observations_note[m][i]:
                            # 计算曲率分母
                            curvature_denominator = distance_value(first_j+1,first_j)+distance_value(first_j,first_j-1)
                            # 计算曲率
                            curvature = (incline - incline_stay) * 2 / curvature_denominator
                            ws_list_att[m][4].cell(first_j - 1, i + 2).value = curvature
                            incline_stay = incline
                            # 增加序号
                            if i == 0:
                                ws_list_att[m][4].cell(first_j - 1, i + 1).value = first_j - 1
                        else:
                            incline_stay = incline  # 用于计算曲率

    wb.save(exist_path)


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    path = "test.xlsx"

    read_excel(path)

    print("No Problem, Sir!!!")

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
